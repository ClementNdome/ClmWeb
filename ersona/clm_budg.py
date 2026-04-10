from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from datetime import datetime

# Function to create the improved Excel file
def create_improved_budget_excel(filename='improved_budget.xlsx'):
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # --- Transactions Sheet ---
    ws_trans = wb.create_sheet('Transactions')
    
    # Headers
    headers = ['Date', 'Month', 'Type', 'Category', 'Subcategory', 'Amount', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws_trans.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Parse and add historical data (hardcoded from provided DOCUMENT)
    transactions = []
    
    # August data
    aug_date = '2024-08-01'  # Approximate dates
    transactions.extend([
        {'date': aug_date, 'month': 'August', 'type': 'Expense', 'cat': 'Food', 'sub': '', 'amt': 3000, 'notes': ''},
        {'date': aug_date, 'month': 'August', 'type': 'Expense', 'cat': 'Rent', 'sub': '', 'amt': 6000, 'notes': ''},
        {'date': aug_date, 'month': 'August', 'type': 'Expense', 'cat': 'House Stuff', 'sub': '', 'amt': 2000, 'notes': 'asso. Stuff to 1 & 2'},
        {'date': aug_date, 'month': 'August', 'type': 'Expense', 'cat': 'Transport', 'sub': '', 'amt': 1680, 'notes': ''},
        {'date': aug_date, 'month': 'August', 'type': 'Expense', 'cat': 'Family & Friends', 'sub': '', 'amt': 1300, 'notes': ''},
        {'date': aug_date, 'month': 'August', 'type': 'Expense', 'cat': 'Personal', 'sub': 'Mistake to GF', 'amt': 1270, 'notes': ''},
        {'date': aug_date, 'month': 'August', 'type': 'Expense', 'cat': 'Personal', 'sub': 'Gifts to GF', 'amt': 460, 'notes': ''},
        {'date': aug_date, 'month': 'August', 'type': 'Expense', 'cat': 'Airtime/Data', 'sub': '', 'amt': 400, 'notes': ''},
        {'date': aug_date, 'month': 'August', 'type': 'Income', 'cat': 'Salary', 'sub': '', 'amt': 20000, 'notes': ''},
        {'date': aug_date, 'month': 'August', 'type': 'Income', 'cat': 'From Family', 'sub': '', 'amt': 500, 'notes': ''},
        {'date': aug_date, 'month': 'August', 'type': 'Savings', 'cat': 'Savings', 'sub': 'August Salary', 'amt': 5000, 'notes': 'To MMF'}
    ])
    
    # September data - more detailed, approximate dates
    sep_date = '2024-09-01'
    transactions.extend([
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Unga 4kg', 'amt': 280, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Mala', 'amt': 200, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Cooking Oil', 'amt': 305, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Onions', 'amt': 150, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Tomatoes', 'amt': 330, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Saumu', 'amt': 50, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Bananas', 'amt': 150, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Kamande', 'amt': 865, 'notes': '#3'},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Eggs', 'amt': 300, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Coffee', 'amt': 114, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Fruits[apples]', 'amt': 100, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Avocado', 'amt': 135, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Potatoes', 'amt': 250, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Milk', 'amt': 150, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Hoho', 'amt': 20, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Meat', 'amt': 150, 'notes': '#frid 12'},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Carrots', 'amt': 70, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Coriadale', 'amt': 20, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Veges', 'amt': 110, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Oranges', 'amt': 50, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Tangawizi and Saumu', 'amt': 40, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Ripe Bananas', 'amt': 60, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Coffee', 'amt': 250, 'notes': '#2'},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Food', 'sub': 'Peanuts', 'amt': 230, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Yummy Hotel', 'amt': 150, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Café', 'amt': 80, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Smk', 'amt': 40, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Lunch 9th', 'amt': 120, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': '11th', 'amt': 20, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': '14th', 'amt': 30, 'notes': '#ngumu'},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': '18th', 'amt': 30, 'notes': '#ngumu'},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': '20', 'amt': 20, 'notes': '#ngumu'},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': '27th', 'amt': 30, 'notes': '#ngumu'},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': '30', 'amt': 30, 'notes': '#ngumu'},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Rent', 'sub': '', 'amt': 6000, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Omo', 'amt': 200, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Bar Soap', 'amt': 200, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Tissues Pair', 'amt': 80, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Dustmat', 'amt': 69, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Mbox', 'amt': 20, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Nivea Cologne', 'amt': 416, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Socks', 'amt': 98, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Super Glue', 'amt': 35, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Transport', 'sub': 'Café Covered', 'amt': 480, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Transport', 'sub': 'Personal', 'amt': 60, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Transport', 'sub': '9th', 'amt': 70, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Transport', 'sub': '11th', 'amt': 20, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Transport', 'sub': '12th', 'amt': 120, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Transport', 'sub': '17th', 'amt': 120, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Transport', 'sub': '22nd', 'amt': 120, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Transport', 'sub': '30th', 'amt': 120, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Family & Friends', 'sub': 'Alfo', 'amt': 500, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Family & Friends', 'sub': 'Mum', 'amt': 1100, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Family & Friends', 'sub': 'Alfo', 'amt': 1000, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Family & Friends', 'sub': 'SpatioAI Domain', 'amt': 500, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Family & Friends', 'sub': 'Rentwater', 'amt': 200, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Family & Friends', 'sub': 'Lucy Fare', 'amt': 150, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Family & Friends', 'sub': 'Poghisho', 'amt': 20, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Airtime/Data', 'sub': 'Wk1', 'amt': 80, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Airtime/Data', 'sub': 'Wk2', 'amt': 40, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Airtime/Data', 'sub': 'Wk3', 'amt': 90, 'notes': '#to refund'},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Airtime/Data', 'sub': 'Wk4', 'amt': 50, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Airtime/Data', 'sub': 'Wk5', 'amt': 20, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'License', 'sub': '', 'amt': 750, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Lamp and Stand', 'sub': '', 'amt': 3000, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Table +3k Trans', 'sub': '', 'amt': 10000, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Expense', 'cat': 'Withdrawal-For Use', 'sub': '', 'amt': 3000, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Income', 'cat': 'Salary', 'sub': '', 'amt': 20000, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Income', 'cat': 'From Café', 'sub': '', 'amt': 1800, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Income', 'cat': 'New Gig from Rachael', 'sub': '', 'amt': 15750, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Savings', 'cat': 'Savings', 'sub': 'Previously on HB', 'amt': 7200, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Savings', 'cat': 'Savings', 'sub': 'August Salary', 'amt': 5000, 'notes': ''},
        {'date': sep_date, 'month': 'September', 'type': 'Savings', 'cat': 'Savings', 'sub': 'September Salary', 'amt': 7500, 'notes': ''}
    ])
    
    # October data
    oct_date = '2024-10-01'
    transactions.extend([
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Rent', 'sub': '', 'amt': 6200, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Unga 2kg', 'amt': 147, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Rice', 'amt': 234, 'notes': '#2kg kheti'},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Onions', 'amt': 70, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Tomatoes', 'amt': 120, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Saumu', 'amt': 25, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Bananas', 'amt': 100, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Kamande', 'amt': 0, 'notes': '#3'},  # Amount not specified, assuming 0 or skip if needed
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Eggs', 'amt': 335, 'notes': '#27'},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Fruits[apples]', 'amt': 0, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Avocado', 'amt': 30, 'notes': '#1'},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Potatoes', 'amt': 0, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Milk', 'amt': 241, 'notes': '2pcs , 3l65'},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Hoho', 'amt': 20, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Meat', 'amt': 200, 'notes': '#sun12'},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Carrots', 'amt': 50, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Coriadale', 'amt': 10, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Veges', 'amt': 50, 'notes': '#cabbage'},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Food', 'sub': 'Ripe Bananas', 'amt': 210, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Ngumu10th', 'amt': 30, 'notes': '#ngumu'},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Ngumu 12th', 'amt': 30, 'notes': '#ngumu'},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Samosa 16th', 'amt': 10, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Hips 21st', 'amt': 100, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Ngumu28th', 'amt': 30, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Ngumu29th', 'amt': 30, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Chicken', 'amt': 100, 'notes': '#150 by robert'},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Shopping Bag', 'amt': 20, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Matchbox', 'amt': 20, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Shaving', 'amt': 100, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Omo', 'amt': 0, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Bar Soap', 'amt': 0, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Tissues Pair', 'amt': 0, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Dustmat', 'amt': 0, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Mbox', 'amt': 0, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Nivea Cologne', 'amt': 0, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Gas Refill', 'amt': 1200, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Matchbox', 'amt': 25, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Panadol', 'amt': 20, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Carrier Bag', 'amt': 10, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Transport', 'sub': '12th', 'amt': 120, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Transport', 'sub': '17th', 'amt': 0, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Transport', 'sub': '22nd', 'amt': 0, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Transport', 'sub': '30th', 'amt': 100, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Family & Friends', 'sub': 'Alfo', 'amt': 250, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Family & Friends', 'sub': 'Mum', 'amt': 1150, 'notes': '#const'},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Family & Friends', 'sub': 'Mum', 'amt': 1200, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Airtime/Data', 'sub': 'Wk1', 'amt': 180, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Airtime/Data', 'sub': 'Wk2', 'amt': 20, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Airtime/Data', 'sub': 'Wk3', 'amt': 40, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Airtime/Data', 'sub': 'Wk4', 'amt': 50, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Expense', 'cat': 'Airtime/Data', 'sub': 'Wk5', 'amt': 20, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Income', 'cat': 'Salary', 'sub': '', 'amt': 20000, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Income', 'cat': 'Domain Reg Main', 'sub': '', 'amt': 100, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Income', 'cat': 'From Clem Waituka', 'sub': '', 'amt': 200, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Income', 'cat': 'Last Ms Mshwari Bal', 'sub': '', 'amt': 900, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Savings', 'cat': 'Savings', 'sub': 'Previous Savings', 'amt': 19450, 'notes': ''},
        {'date': oct_date, 'month': 'October', 'type': 'Savings', 'cat': 'Savings', 'sub': 'On Oct', 'amt': 7500, 'notes': ''}
    ])
    
    # November data
    nov_date = '2024-11-01'
    transactions.extend([
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Rent', 'sub': '', 'amt': 6200, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Food', 'sub': 'Unga 2kg', 'amt': 148, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Food', 'sub': 'Rice', 'amt': 214, 'notes': '#1kg kheti'},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Food', 'sub': 'Onions', 'amt': 50, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Food', 'sub': 'Tomatoes', 'amt': 50, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Food', 'sub': 'Saumu', 'amt': 50, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Food', 'sub': 'Eggs', 'amt': 220, 'notes': '#27'},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Food', 'sub': 'Milk', 'amt': 118, 'notes': '2pcs , 1l65'},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Food', 'sub': 'Veges', 'amt': 0, 'notes': '#cabbage'},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Ngumu10th', 'amt': 0, 'notes': '#ngumu'},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Ngumu 12th', 'amt': 0, 'notes': '#ngumu'},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Samosa 16th', 'amt': 0, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Hips 21st', 'amt': 0, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Ngumu28th', 'amt': 0, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Eat Outs', 'sub': 'Ngumu29th', 'amt': 0, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Shopping Bag', 'amt': 20, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Matchbox', 'amt': 20, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Shaving', 'amt': 100, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Bathing Soap 3', 'amt': 150, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Steel Wool', 'amt': 20, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Tissue', 'amt': 76, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Tape', 'amt': 27, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'House Stuff', 'sub': 'Carrier Bag', 'amt': 40, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Transport', 'sub': '1st', 'amt': 120, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Transport', 'sub': '17th', 'amt': 0, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Transport', 'sub': '22nd', 'amt': 0, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Transport', 'sub': '30th', 'amt': 0, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Family & Friends', 'sub': 'Alfo', 'amt': 0, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Family & Friends', 'sub': 'Mum', 'amt': 1200, 'notes': '#const'},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Family & Friends', 'sub': 'Lnr', 'amt': 1800, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Expense', 'cat': 'Airtime/Data', 'sub': 'Wk1', 'amt': 40, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Income', 'cat': 'Salary', 'sub': '', 'amt': 20000, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Income', 'cat': 'Prev Months Mshwari Bal', 'sub': '', 'amt': 765, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Income', 'cat': 'Pochi Prev Month Bal', 'sub': '', 'amt': 658, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Income', 'cat': 'Mpesa Bal Rems', 'sub': '', 'amt': 500, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Savings', 'cat': 'Savings', 'sub': 'Previous Savings', 'amt': 26950, 'notes': ''},
        {'date': nov_date, 'month': 'November', 'type': 'Savings', 'cat': 'Savings', 'sub': 'On November', 'amt': 7500, 'notes': ''}
    ])
    
    # Write transactions to sheet
    for row, tx in enumerate(transactions, 2):
        ws_trans.cell(row, 1, tx['date'])
        ws_trans.cell(row, 2, tx['month'])
        ws_trans.cell(row, 3, tx['type'])
        ws_trans.cell(row, 4, tx['cat'])
        ws_trans.cell(row, 5, tx.get('sub', ''))
        ws_trans.cell(row, 6, tx['amt'])
        ws_trans.cell(row, 7, tx['notes'])
    
    # Auto-fit columns
    for col in range(1, 8):
        ws_trans.column_dimensions[get_column_letter(col)].width = 15
    
    # --- Budget Sheet ---
    ws_budget = wb.create_sheet('Budget')
    
    # Categories list (refined)
    categories = [
        'Food', 'Rent', 'House Stuff', 'Transport', 'Family & Friends', 'Personal', 'Airtime/Data', 'Eat Outs', 'License', 'Other'
    ]
    
    # Headers
    budget_headers = ['Category', 'Budgeted', 'Actual', 'Variance']
    for col, header in enumerate(budget_headers, 1):
        cell = ws_budget.cell(1, col, header)
        cell.font = Font(bold=True)
    
    # Populate categories and sample budgets (user can adjust)
    for row, cat in enumerate(categories, 2):
        ws_budget.cell(row, 1, cat)
        ws_budget.cell(row, 2, 0)  # Placeholder for budgeted
        # Formula for actual: SUMIF from Transactions where Type=Expense and Category=cat
        actual_formula = f'=SUMIF(Transactions!D:D, A{row}, Transactions!F:F)'
        ws_budget.cell(row, 3).value = actual_formula
        # Variance = Budgeted - Actual
        variance_formula = f'=B{row}-C{row}'
        ws_budget.cell(row, 4).value = variance_formula
    
    # Total row
    total_row = len(categories) + 2
    ws_budget.cell(total_row, 1, 'Total')
    for col in [2,3,4]:
        ws_budget.cell(total_row, col).value = f'=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row-1})'
    
    # --- Monthly Summary Sheet ---
    ws_summary = wb.create_sheet('Monthly Summary')
    
    # Months from data
    months = ['August', 'September', 'October', 'November']  # Add more as needed
    summary_headers = ['Month', 'Total Income', 'Total Expenses', 'Net', 'Savings Contribution']
    for col, header in enumerate(summary_headers, 1):
        ws_summary.cell(1, col, header)
    
    for row, month in enumerate(months, 2):
        ws_summary.cell(row, 1, month)
        # Income: SUMIF Month and Type=Income
        income_formula = f'=SUMIFS(Transactions!F:F, Transactions!B:B, A{row}, Transactions!C:C, "Income")'
        ws_summary.cell(row, 2).value = income_formula
        # Expenses: SUMIFS Month and Type=Expense
        exp_formula = f'=SUMIFS(Transactions!F:F, Transactions!B:B, A{row}, Transactions!C:C, "Expense")'
        ws_summary.cell(row, 3).value = exp_formula
        # Net = Income - Expenses
        net_formula = f'=B{row}-C{row}'
        ws_summary.cell(row, 4).value = net_formula
        # Savings: SUMIFS Month and Type=Savings
        sav_formula = f'=SUMIFS(Transactions!F:F, Transactions!B:B, A{row}, Transactions!C:C, "Savings")'
        ws_summary.cell(row, 5).value = sav_formula
    
    # --- Savings & Investments Sheet ---
    ws_invest = wb.create_sheet('Savings & Investments')
    
    # Headers
    invest_headers = ['Date', 'Description', 'Contribution', 'Return Rate (%)', 'Projected Return', 'Balance']
    for col, header in enumerate(invest_headers, 1):
        ws_invest.cell(1, col, header)
    
    # Sample data based on user (current MMF, suggest better)
    invest_data = [
        {'date': '2024-08-01', 'desc': 'M-Pesa MMF August', 'contrib': 5000, 'rate': 5, 'proj': '', 'bal': ''},
        {'date': '2024-09-01', 'desc': 'M-Pesa MMF September', 'contrib': 7500, 'rate': 5, 'proj': '', 'bal': ''},
        {'date': '2024-10-01', 'desc': 'M-Pesa MMF October', 'contrib': 7500, 'rate': 5, 'proj': '', 'bal': ''},
        {'date': '2024-11-01', 'desc': 'M-Pesa MMF November', 'contrib': 7500, 'rate': 5, 'proj': '', 'bal': ''},
        # Suggestions
        {'date': datetime.now().strftime('%Y-%m-%d'), 'desc': 'Switch to Gulfcap MMF (12.9% p.a.)', 'contrib': 0, 'rate': 12.9, 'proj': '', 'bal': ''},
        {'date': datetime.now().strftime('%Y-%m-%d'), 'desc': 'Treasury Bills (10%+)', 'contrib': 0, 'rate': 10, 'proj': '', 'bal': ''},
        {'date': datetime.now().strftime('%Y-%m-%d'), 'desc': 'SACCO Account (8-10%)', 'contrib': 0, 'rate': 9, 'proj': '', 'bal': ''},
    ]
    
    for row, item in enumerate(invest_data, 2):
        ws_invest.cell(row, 1, item['date'])
        ws_invest.cell(row, 2, item['desc'])
        ws_invest.cell(row, 3, item['contrib'])
        ws_invest.cell(row, 4, item['rate'])
        # Projected Return = Contribution * (Rate/100) /12 for monthly
        if item['contrib'] > 0:
            proj_formula = f'=C{row} * (D{row}/100) / 12'
            ws_invest.cell(row, 5).value = proj_formula
        # Balance running total
        if row == 2:
            bal_formula = '=C2 + E2'
        else:
            bal_formula = f'=F{row-1} + C{row} + E{row}'
        ws_invest.cell(row, 6).value = bal_formula
    
    # Notes on better investments
    ws_invest.cell(row + 2, 1, 'Notes: Upgrade from 4-6% M-Pesa to higher yield options like MMFs (up to 13%), T-Bills, SACCOs for better returns. Research and diversify.')
    
    # --- Goals Sheet ---
    ws_goals = wb.create_sheet('Goals')
    
    goals_headers = ['Goal', 'Cost', 'Status', 'Deadline', 'Funded', 'Remaining']
    for col, header in enumerate(goals_headers, 1):
        ws_goals.cell(1, col, header)
    
    goals_data = [
        {'goal': 'Pay for License', 'cost': 750, 'status': 'Done', 'deadline': '', 'funded': 750, 'rem': ''},
        {'goal': 'Buy Lamp Light', 'cost': 2400, 'status': 'Done', 'deadline': '', 'funded': 2400, 'rem': ''},
        {'goal': 'Study/Computer Table', 'cost': 7000, 'status': 'Done', 'deadline': '', 'funded': 7000, 'rem': ''},
        {'goal': 'Monitor', 'cost': 30000, 'status': 'Pending', 'deadline': '2025-01-01', 'funded': 0, 'rem': ''},
        {'goal': 'Chair', 'cost': 8500, 'status': 'Pending', 'deadline': '2025-02-01', 'funded': 0, 'rem': ''},
        {'goal': 'Keyboard', 'cost': 3000, 'status': 'Pending', 'deadline': '2025-03-01', 'funded': 0, 'rem': ''},
        {'goal': 'Powerful Laptop', 'cost': 100000, 'status': 'Pending', 'deadline': '2025-03-01', 'funded': 0, 'rem': ''},
        {'goal': 'Get 4 Certifications', 'cost': 0, 'status': 'Pending', 'deadline': '2025-12-31', 'funded': 0, 'rem': ''},
    ]
    
    for row, goal in enumerate(goals_data, 2):
        ws_goals.cell(row, 1, goal['goal'])
        ws_goals.cell(row, 2, goal['cost'])
        ws_goals.cell(row, 3, goal['status'])
        ws_goals.cell(row, 4, goal['deadline'])
        ws_goals.cell(row, 5, goal['funded'])
        rem_formula = f'=B{row}-E{row}'
        ws_goals.cell(row, 6).value = rem_formula
    
    # Save the workbook
    wb.save(filename)
    print(f'Excel file "{filename}" created successfully!')

# Run the function
create_improved_budget_excel()